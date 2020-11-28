import openpyxl
wb = openpyxl.load_workbook(filename= './Baza_2019-2020.xlsx')
sheet = wb.get_sheet_by_name('Проживающие')
sheet2 = wb.get_sheet_by_name('Вахта')

def read_write():
    count = i = 2
    k = 1

    while sheet.cell(row= i, column= 1).value != None:

        if (sheet.cell(row= i, column= 2).value != 'женское' and
            sheet.cell(row= i, column= 2).value !='занято' and
            sheet.cell(row= i, column= 2).value != 'мужское'):

            if count % 2 == 0:
                sheet2.cell(row=k, column= 1).value = sheet.cell(row= i, column= 1).value
                sheet2.cell(row=k, column= 2).value = sheet.cell(row= i, column= 2).value
            else:
                sheet2.cell(row= k, column= 3).value = sheet.cell(row= i, column= 1).value
                sheet2.cell(row= k, column= 4).value = sheet.cell(row=i, column= 2).value
                k += 1
            i += 1
            count += 1
        else:
            i += 1
    wb.save('./Output.xlsx')

read_write()
